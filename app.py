"""
   author: josuerom
   date:   12/10/2025 16:34:05
"""
from flask import Flask, render_template, request, session, send_file, redirect, url_for
from openpyxl import load_workbook
import os
from datetime import datetime
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)


CHECKLISTS = {
   'pc': {
      'titulo': 'Checklist Proquinal PC 2025',
      'empresa': 'Proquinal',
      'tipo': 'PC',
      'plantilla': 'plantilla_pc.xlsx',
      'preguntas': [
         'CON USUARIO LOCAL SOPORTE. El usuario de configuración inicial se crea como: "Soporte" y se asigna contraseña.',
         'Realizar Windows Update.',
         'Realizar actualización de drivers en el equipo.',
         'Habilitar usuario administrador y cambiar nombre a: "Adminpc_pqn" (no deshabilitar usuario "Soporte").',
         'Verificar activación de licencia de Windows.',
         'Instalar .NET Framework 3.5 desde la página de Microsoft.',
         'Activar características adicionales de Framework 3.5 (desde appwiz.cpl).',
         'Particionar el disco duro (si es posible): disco de 500 GB (250 para D: y el resto para C:), disco de 1 TB (600 para D: y el resto para C:) [la unidad D: se debe llamar "Datos"].',
         'Agregar o quitar programas.',
         'Subir el equipo al dominio (proquinal.com). El nombre del equipo debe contener el service tag. Ejemplo: "CBY45TF-PQN-COL" [en equipos Lenovo usar los últimos 7 caracteres del serial].',
         'En la descripción del equipo colocar: "Activo" (con A mayúscula), el número de activo, y el correo. Ejemplo: "Activo 35987, hola.pqn@spradling.group".',
         'Instalar Oracle (copiar carpetas ORAWIN95 y ORA9i, importar registro).',
         'Borrar carpetas anteriores de Oracle (ORANT.OLD, ORAWIN95.OLD, ORAPQN.OLD, ORACLEXE.OLD, ORA9I.OLD).',
         'Crear la carpeta C:\\ Listados.',
         'Instalar Acrobat Reader desde internet (última versión).',
         'Instalar última versión de Java (versión 8-341, última compatible con Oracle).',
         'Desactivar actualización automática de Java.',
         'Copiar la carpeta ORAPQN en C:\\ ORAPQN y el backup de TURIN en C:\\ WINDOWS.',
         'Configurar archivo "hosts" (si es portátil): copiar archivo desde recursos/instaladores/host [es para OracleWeb y Avaya].',
         'Configurar variable PATH para ingreso a Oracle Web (C:\\ ORAPQN\\ LIB60) y para Innovación.',
         'Configurar variable de entorno TNS_ADMIN con valor "O:\\ tnsnames".',
         'Agregar en propiedades de red: DNS 172.16.1.1 como principal y 172.16.1.32 como alternativo (solo para equipos de escritorio dentro de PQN).',
         'CON USUARIO DE DOMINIO. Cambiar configuración de energía del equipo (colocar "Nunca" en opciones de energía; batería solo 3 horas; desactivar inicio rápido).',
         'Redireccionar "Mis documentos" a la carpeta D:\\ Datos [la carpeta debe llamarse "DATOS", no solo la unidad D:].',
         'Enrolar el equipo (agregar cuenta de correo en "Acceder al trabajo o colegio"). Validar que esté en el grupo de equipos Proquinal del AD.',
         'Instalar Office (validar versión y licencia del usuario).',
         'Configurar el correo con archivo OST en D:\\ Correo (No Borrar). Enviar formato de firma para que el usuario la configure.',
         'Instalar Teams (habilitar inicio en segundo plano; desactivar inicio automático con Windows).',
         'Habilitar uso de OneDrive del usuario (si tiene licencia). Ruta: D:\\ . ¡¡¡No sincronizar OneDrive con escritorio, documentos ni imágenes!!!',
         'Deshabilitar opción "Ahorrar espacio y descargar los archivos cuando los use".',
         'Habilitar conexión a Escritorio Remoto con el usuario del equipo.',
         'Habilitar macros en Excel y agregar a TURIN (file://172.16.1.22) como sitio de confianza.',
         'Predeterminar Acrobat Reader para abrir archivos PDF.',
         'Habilitar asistencia remota (instalar TeamViewer Host 15) y registrarlo en la consola de administración.',
         'Reconectar unidades de red (carpetas compartidas).',
         'Instalar fuentes corporativas.',
         'Instalar Citrix Online (última versión, solo si es portátil).',
         'Instalar software VPN Fortinet (última versión, solo si es portátil).',
         'Instalar acceso a backup verificando que exista la carpeta en TURIN.',
         'Copiar accesos directos de aplicaciones Colombia y Costa Rica (crear .BAT en ícono de Oracle para conexión a O:\\).',
         'Copiar accesos directos de aplicaciones Office.',
         'Instalar Microsoft Edge y Google Chrome.',
         'Configurar Edge para visualizar páginas en modo Internet Explorer. Agregar links de Oracle Web a la lista de permitidos (modo de compatibilidad IE).',
         'Configurar seguridad de Java para acceso a Oracle Web.',
         'Cambiar configuración regional para separación de miles (",") y decimales (".").',
         'Crear acceso directo a backup TURIN.',
         'Crear acceso directo a Daruma: https://proquinal.darumasoftware.com/app.php/staff/ (dejar como favorito en el navegador).',
         'Crear acceso directo a Mayte: https://mayte.spradling.group/ (dejar como favorito en el navegador).',
         'Instalar impresoras del área correspondiente.',
         'Dejar predeterminada la página https://spradling.group/es-la en todos los navegadores.',
         'Instalar módulo de Acronis Backup y registrarlo en la consola (crear carpeta para backup en la ruta D:\\ Datos\\ [año en curso]).',
         'Aplicar política de backup 2023 en la consola de Acronis (confirmar que quede funcionando).',
         'Realizar backup y/o copia de datos: copiar archivos de favoritos y escritorio desde Documents and Settings.',
         'Verificar que el software instalado en el equipo tenga licencia.',
         'Verificar si el usuario utiliza otro tipo de software (como IRIS, archivo de planeación, software de diseño, etc.).',
         'Habilitar contraseña de arranque (solo si es portátil).',
         'Validar activación de MFA en el correo del usuario.',
         'Habilitar contraseña de ingreso a la BIOS y de encendido del equipo (solo si es portátil).',
         'Verificar que el equipo quede con actualizaciones al día (Windows, Office, etc.).',
         'Habilitar punto de restauración automática para la unidad C: y crear un punto de restauración.',
         'CON USUARIO LOCAL ADMINPC_PQN. Subir el equipo a Windows 11.',
         'Asignar activo fijo.',
         'Realizar inventario en el sistema: físico (incluir número de RAF y activos que se cambian) y digital en D:\\ Imagen (No Borrar).',
         'Diligenciar datos en Oracle del activo correspondiente. Equipo: 1, 2, 4, 5, 6, 7, 8, 79, 300, 331 (nombre de quien realiza el inventario); 339, 385, 413, 414, 415, 449, 451, 464 (código del empleado, solo números). Monitor: 1, 2, 4, 13, 15, 77, 79, 331, 413, 415, 453, 464.',
         'Diligenciar checklist en D:\\ Imagen (No Borrar).',
         'Crear imagen de la partición C:[SISTEMAPQN]. Crear archivo README informativo en D:\\ Imagen (No Borrar). ¡¡¡Solo equipos especiales!!!',
         'Revisar que se pueda ver el equipo en el servidor de Intune y Defender.',
         'Confirmar borrado del equipo antiguo en Intune y Defender (para liberar licencias).',
         'Verificar que los archivos en la unidad D:\\ Datos (estén cifrados).',
         'Verificar que no haya software en la unidad de CD.',
         'Realizar acta de entrega del equipo.',
         'Realizar traslado del equipo antiguo al área de soporte.',
         'Mantenimiento preventivo de CPU.',
         'Limpieza de pantalla.',
         'Realizar el traslado del equipo viejo y dejarlo marcado con usuario, fecha y responsable.'
      ]
   },
   'terminales': {
      'titulo': 'Checklist Proquinal Terminales 2025',
      'empresa': 'Proquinal',
      'tipo': 'Terminales',
      'plantilla': 'plantilla_terminales.xlsx',
      'preguntas': [
         'CON USUARIO SOPORTE. El usuario de configuración inicial se crea como: "Soporte" y se asigna contraseña.',
         'Realizar Windows Update.',
         'Realizar actualización de drivers en el equipo.',
         'Habilitar usuario administrador y cambiar el nombre a: "Adminpc_pqn" (no olvidar deshabilitar el usuario "Soporte").',
         'Verificar activación de la licencia de Windows.',
         'Instalar .NET Framework 3.5 desde la página de Microsoft.',
         'Activar características adicionales de Framework 3.5 (desde appwiz.cpl).',
         'Agregar o quitar programas.',
         'Subir el equipo al dominio (proquinal.com). El nombre del equipo debe contener el Service Tag. Ejemplo: "CBY45TF-PQN-COL" [equipos Lenovo: usar los últimos 7 caracteres del serial].',
         'En la descripción del equipo, colocar: "Activo" (con A mayúscula), espacio, número de activo y correo. Ejemplo: "Activo 17800, user@spradling.group".',
         'Instalar Oracle (copiar carpetas ORAWIN95 y ORA9i, importar registro).',
         'Borrar carpetas anteriores de Oracle: ORANT.OLD, ORAWIN95.OLD, ORAPQN.OLD, ORACLEXE.OLD, ORA9I.OLD.',
         'Crear la carpeta "C:\\ Listados".',
         'Instalar Acrobat Reader desde internet (última versión).',
         'Instalar última versión de Java (versión 8-341, última compatible con Oracle).',
         'Desactivar actualización automática de Java.',
         'Copiar la carpeta ORAPQN en "C:\\ ORAPQN" y el backup de TURIN en "C:\\ WINDOWS".',
         'Configuración de la variable PATH para ingreso a Oracle Web ("C:\\ ORAPQN\\ LIB60") e Innovación.',
         'Configuración de la variable de entorno "TNS_ADMIN" con valor: "O:\\ tnsnames".',
         'Agregar en propiedades de red: DNS 172.16.1.1 como principal y 172.16.1.32 como alternativo.',
         'Verificar por Regedit que aparezca la variable SCADA: nombre de cadena: "Path_Scada"; valor: "\\\\172.16.1.96\\scada$".',
         'Verificar por Regedit que aparezca la variable para Porterías (si aplica): nombre de cadena: "Path_Visitas"; valor: "C:\\ fotos_visitas".',
         'CON USUARIO DE DOMINIO. Cambiar la configuración de energía del equipo: colocar "Nunca" en las opciones de energía, batería con solo 3 horas, y desactivar el inicio rápido.',
         'Enrolar el equipo (agregar cuenta de correo en "Acceder al trabajo o colegio"). Validar que esté en el grupo de equipos Proquinal del AD.',
         'Habilitar conexión a Escritorio Remoto con el usuario del equipo.',
         'Colocar a TURIN ("file://172.16.1.22") como sitio de confianza desde Opciones de Internet.',
         'Predeterminar Acrobat Reader para abrir archivos PDF.',
         'Habilitar asistencia remota (instalar TeamViewer Host Customizado 15) y registrarlo en la consola de administración (validar que tenga la extensión "911").',
         'Crear archivo .BAT para reconexión con la carpeta: "O:\\172.16.1.22\\ orapqn$" en caso de desconexión.',
         'Instalar Citrix Online (última versión).',
         'Copiar accesos directos de aplicaciones Colombia y Costa Rica (crear .BAT en el ícono de Oracle para conexión a "O:\\").',
         'Instalar visores de Office.',
         'Instalar Microsoft Edge y Google Chrome.',
         'Configurar Edge para visualización de páginas en modo Internet Explorer. Agregar link(s) de Oracle Web a la lista de sitios permitidos (compatibilidad IE).',
         'Configurar seguridad de Java para acceso a Oracle Web.',
         'Cambiar configuración regional para separación de miles (",") y decimales (".").',
         'Crear acceso directo a Daruma: https://proquinal.darumasoftware.com/app.php/staff/ (dejar como favorito en el navegador).',
         'Crear acceso directo a Mayte: https://mayte.spradling.group/ (dejar como favorito en el navegador).',
         'Crear acceso directo a la página de Sustancias Químicas: https://spradling.aybapps.net/PROQUINAL.SUSTANCIAS/(S(bfumxph5htfsk0jv0dpbke5k))/InicioSustancias.aspx.',
         'Configurar acceso a SCADA desde navegador: "\\\\172.16.1.96\\".',
         'Instalar impresoras del área correspondiente (si aplica).',
         'Dejar predeterminada la página: https://spradling.group/es-la en todos los navegadores.',
         'Configurar AutoLogon del usuario desde Regedit: modificar "AutoAdminLogon" (cambiar valor de 0 a 1); "DefaultUserName" (usuario de la terminal con dominio);',
         'Crear clave: "DefaultPassword" con el valor correspondiente (contraseña del usuario).',
         'Verificar que el equipo tenga todas las actualizaciones al día (Windows, Office, etc.).',
         'Habilitar punto de restauración automática para la unidad C: y crear un punto de restauración.',
         'Subir el equipo a Windows 11.',
         'Asignar Activo Fijo.',
         'Realizar inventario en el sistema: físico (incluir número de RAF y activos que se cambian) y digital (guardar en "D:\\ Imagen (No Borrar)).',
         'Diligenciar los datos en Oracle del activo correspondiente. Equipo: 1, 2, 4, 5, 6, 7, 8, 79, 300, 331 (nombre de quien realiza el inventario); 339, 385, 413, 414, 415, 449, 451, 464 (código del empleado, solo números). Monitor: 1, 2, 4, 13, 15, 77, 79, 331, 413, 415, 453, 464.',
         'Diligenciar check list en "D:\\ Imagen (No Borrar).',
         'Revisar que el equipo se visualice en el servidor de Intune y Defender.',
         'Confirmar borrado del equipo antiguo en Intune y Defender (para liberar licencias).',
         'Realizar acta de entrega del equipo.',
         'Realizar traslado del equipo antiguo al área de soporte.',
         'Mantenimiento preventivo de CPU.',
         'Limpieza de pantalla.',
         'Realizar traslado del equipo viejo y dejarlo marcado con: usuario, fecha y responsable.'
      ]
   },
   'macos': {
      'titulo': 'Checklist Proquinal MacOS 2025',
      'empresa': 'Proquinal',
      'tipo': 'MacOS',
      'plantilla': 'plantilla_macos.xlsx',
      'preguntas': [
         'CON USUARIO ADMINPC_PQN. El usuario de configuración inicial se crea como: "Adminpc_pqn" y se asigna contraseña.',
         'Realizar actualizaciones de MacOS.',
         'Instalar TeamViewer Host (última versión) en la máquina y registrarlo en la consola de administración.',
         'Crear usuario correspondiente al propietario de la máquina.',
         'CON USUARIO DE DOMINIO. Ingresar con el usuario de la máquina.',
         'Realizar la instalación del portal de Microsoft para enrolamiento por Intune.',
         'Validar correcto registro en la consola de Intune y Defender.',
         'Instalar licencia de Office correspondiente a la licencia asignada al usuario de la máquina.',
         'Configurar la suite de Office con la cuenta del usuario.',
         'Configurar OneDrive en la máquina (validar si el usuario desea que los archivos se descarguen o se mantengan en la nube).',
         'Instalar Microsoft Teams. Dar permisos para compartir pantalla y audio.',
         'Instalar Adobe Cloud (si el usuario tiene licencia).',
         'Validar con el usuario qué aplicativos de la suite de Adobe necesita e instalar los requeridos.',
         'Instalar Google Chrome y FortiClient VPN.',
         'Predeterminar Acrobat Reader para abrir archivos PDF.',
         'Validar con el área de Infraestructura los permisos de navegación para el equipo.',
         'Realizar el paso de datos del equipo anterior a la nueva máquina.',
         'Revisar que el equipo tenga instalada la última versión de macOS (Monterey, Ventura o Sonoma; priorizar la más reciente).',
         'Configurar hardware externo que utilice el usuario (tableta de dibujo, graphic pen, etc.).',
         'Configurar FileVault: 1) Abrir Preferencias del Sistema. 2) Seleccionar "Seguridad y privacidad". 3) Seleccionar "FileVault". 4) Seleccionar "Activar FileVault".'
      ]
   },
   'tablets': {
      'titulo': 'Checklist Proquinal Tablets 2025',
      'empresa': 'Proquinal',
      'tipo': 'Tablets',
      'plantilla': 'plantilla_tablets.xlsx',
      'preguntas': [
         'Antes de iniciar sesión, enrolar en Intune escaneando el código QR.',
         'La tablet aparece en Intune al escanear el código QR.',
         'Loguear la cuenta de Play Store corporativa: "soportesistemaspqncol@gmail.com".',
         'Iniciar sesión en Office 365 con la cuenta corporativa del usuario.',
         'Habilitar el inicio de sesión por PIN (4 dígitos) y tomar nota para entregar al usuario.',
         'Habilitar el bloqueo de la tablet para 15 minutos (no menos).',
         'Instalar Outlook.',
         'Instalar Office (Word, Excel, PowerPoint).',
         'Instalar Power BI.',
         'Instalar OneDrive.',
         'Instalar TeamViewer Host.',
         'Instalar Adobe Reader.',
         'Instalar la app "HubSpot" desde Play Store.',
         'Crear acceso directo a Daruma: https://proquinal.darumasoftware.com/app.php/staff/ (dejar como favorito en el navegador).',
         'Crear acceso directo a Mayte: https://mayte.spradling.group/ (dejar como favorito en el navegador).',
         'Crear acceso directo a la Biblioteca de Productis: https://library.spradling.group/auth/login (dejar como favorito en el navegador).',
         'Dejar como predeterminada la página https://spradling.group/en-la en todos los navegadores.',
         'Verificar si el usuario utiliza otro tipo de software como IRIS, archivo de planeación, software de diseño, etc.',
         'Verificar que el software instalado en el equipo cuente con licencia.',
         'Dejar la tablet con las últimas actualizaciones disponibles.',
         'Asignar activo fijo.',
         'Realizar inventario en el sistema: 1) Físico (incluir número de RAF y activos que se cambian). 2) Digital en "D:\\ Imagen (No Borrar).',
         '23. Diligenciar los datos en Oracle del activo correspondiente: Equipo: 1 - 2 - 4 - 5 - 6 - 7 - 8 - 79 - 300 - 331 - 339 - 385 - 413 - 414 - 415 - 449 - 451 - 464. Monitor: 1 - 2 - 4 - 13 - 15 - 77 - 79 - 331 - 413 - 415 - 453 - 464.',
         'Diligenciar checklist en "D:/Imagen (No Borrar).',
         'Revisar que se pueda ver en el servidor de Defender e Intune.',
         'Realizar el traslado del equipo viejo y dejarlo marcado con usuario, fecha y responsable.'
      ]
   },
   'calypso': {
      'titulo': 'Checklist Calypso CCS CBQ 2025',
      'empresa': 'Calypso',
      'tipo': 'CCS-CBQ',
      'plantilla': 'plantilla_calypso.xlsx',
      'preguntas': [
         "DEBE USAR AUTOPILOT. Iniciar sesión con el equipo hasta el momento que pide internet",
         "Enrolar dentro del tenant de Intune",
         "Validar que el equipo se pueda ver desde Entra",
         "Asignar el usuario en la inscripción de equipos por Autopilot",
         "Agregar el equipo al grupo de Intune: GRP_Dispositivos Windows Autopilot CCS-COL (para la instalación de aplicaciones)",
         "Confirmar que el equipo aparezca enrolado en Entra y ajustar las etiquetas: tipo de dispositivo, etiqueta de empresa, grupos asignados (Equipos CCS-COL + el de despliegue de apps)",
         "Iniciar sesión con el usuario asignado en el equipo (la sesión se inicia con el correo electrónico)",
         "Habilitar usuario administrador \"Adminpc_ccs\"",
         "Verificar activación de licencia de Windows",
         "Instalar .NET Framework 3.5",
         "Realizar Windows Update",
         "Verificar controladores del equipo (actualización de drivers)",
         "Agregar o quitar programas (eliminar todo lo relacionado con apps de juegos y redes sociales); dejar solo apps corporativas",
         "Particionar el disco duro (si es posible): disco 500 GB (250 GB para D: y el resto para C:), disco 1 TB (600 GB para D: y el resto para C:)",
         "Instalar Comerssia (si lo requiere) ***solo PDV*** (tener presente que la hora regional cambia)",
         "En el grupo de trabajo del equipo, colocar la ciudad o sede a la que pertenece",
         "En la descripción del equipo, colocar \"Activo [número de activo], correo@spradling.group\" (Ej.: Activo 175608, hola.pqn@spradling.group)",
         "Cambiar el nombre del equipo (debe contener el Service Tag). Ej.: P08DTGF-[CCS-CBQ]-COL. *Equipos Lenovo: usar los últimos 7 caracteres del serial*. Colocar grupo de trabajo con usuario",
         "CON USUARIO DE AZUREAD: Cambiar la configuración de energía del equipo (colocar 'nunca' en las opciones de energía. En batería: solo 3 horas)",
         "Direccionar 'Mis documentos' a la carpeta D:\\ Datos",
         "Instalar Office según licencia del usuario (validar ejecución de tarea desde Intune)",
         "Instalar Acrobat Reader (validar ejecución de tarea desde Intune)",
         "Instalar Chrome Enterprise (validar ejecución de tarea desde Intune)",
         "Configurar el correo con el archivo OST en D:\\ Datos\\ Correo",
         "Instalar Teams (habilitar el inicio en segundo plano, no al inicio con Windows)",
         "Habilitar uso de OneDrive del usuario (si tiene licencia); guardar ubicación en D:",
         "Deshabilitar opción 'Ahorrar espacio y descargar los archivos cuando los use'; no sincronizar Documentos, Escritorio ni Imágenes",
         "Predeterminar Acrobat Reader para abrir archivos PDF",
         "Validar que el equipo tome el fondo de pantalla según la política de Intune",
         "Habilitar asistencia remota (validar tarea desde Intune: TeamViewer Host 15) y registrarlo en la consola de administración",
         "Instalar fuentes corporativas",
         "Copiar accesos directos a las aplicaciones de Office",
         "Cambiar la configuración regional para separación de miles y decimales ('.' para decimales, ',' para miles)",
         "Crear acceso directo en el escritorio a SAP (mstsc -v sapcalypso.spradling.group:9546)",
         "Instalar software VPN Fortinet (solo instalar, no configurar)",
         "Dejar como página predeterminada https://www.tiendascalypso.com en todos los navegadores",
         "Crear acceso directo a MAYTE: https://mayte.spradling.group/es-la (dejar como favorito en navegador)",
         "Habilitar conexión a escritorio remoto con el usuario del equipo",
         "Validar activación de MFA en el correo del usuario",
         "Habilitar contraseña de ingreso al sistema operativo (si es portátil)",
         "Verificar si el usuario utiliza otro tipo de software como Adobe, AutoCAD, software de diseño, etc.",
         "Verificar que el equipo quede con actualizaciones al día (Windows, Office, etc.)",
         "Asignar activo fijo (AF)",
         "Habilitar punto de restauración automática para la unidad C:",
         "Realizar inventario en el sistema (físico y digital en D:\\ Imagen (No Borrar)",
         "Diligenciar check list en D:\\ Imagen (No Borrar)",
         "Crear imagen de la partición C:\\ [SISTEMAPQN] y readme informativo en D:\\ Imagen (No Borrar) ¡¡¡solo equipos especiales!!!",
         "Revisar que se pueda ver en el servidor de antivirus Defender",
         "Revisar que el equipo se vea desde Intune y Entra sin estar duplicado; retirar los equipos antiguos de estas plataformas"
      ]
   }
}


@app.route('/')
def index():
   return render_template('index.html', checklists=CHECKLISTS)


@app.route('/formulario/<tipo>', methods=['GET', 'POST'])
def formulario(tipo):
   if tipo not in CHECKLISTS:
      return redirect(url_for('index'))
   
   if request.method == 'POST':
      session['activo_fijo'] = request.form.get('activo_fijo')
      session['propietario'] = request.form.get('propietario')
      session['cargo'] = request.form.get('cargo')
      session['tecnico'] = request.form.get('tecnico')
      session['tipo_checklist'] = tipo
      return redirect(url_for('checklist', tipo=tipo))
   
   return render_template('checklist_form.html', 
                        checklist=CHECKLISTS[tipo],
                        tipo=tipo)


@app.route('/checklist/<tipo>', methods=['GET', 'POST'])
def checklist(tipo):
   if tipo not in CHECKLISTS:
      return redirect(url_for('index'))
   
   if 'activo_fijo' not in session:
      return redirect(url_for('formulario', tipo=tipo))
   
   if request.method == 'POST':
      respuestas = {}
      for i in range(len(CHECKLISTS[tipo]['preguntas'])):
         respuestas[i+1] = request.form.get(f'pregunta_{i+1}', 'N/A')
      
      archivo_generado = generar_excel(tipo, respuestas)
      
      return send_file(
         archivo_generado,
         as_attachment=True,
         download_name=os.path.basename(archivo_generado)
      )
   
   # CORRECCIÓN: Pasar las preguntas correctamente al template
   preguntas_enumeradas = list(enumerate(CHECKLISTS[tipo]['preguntas'], 1))
   
   return render_template('checklist.html',
                        checklist=CHECKLISTS[tipo],
                        tipo=tipo,
                        preguntas=preguntas_enumeradas,
                        session=session)


def generar_excel(tipo, respuestas):
   config = CHECKLISTS[tipo]
   plantilla_path = os.path.join('templates_excel', config['plantilla'])
   
   # Cargar plantilla existente
   wb = load_workbook(plantilla_path)
   ws = wb.active
   
   # Buscar y llenar respuestas en columna C según ID en columna A
   line = None
   for row in range(1, ws.max_row + 1):
      celda_id = ws.cell(row=row, column=1).value
      if celda_id and str(celda_id).isdigit():
         pregunta_id = int(celda_id)
         if pregunta_id in respuestas:
            ws.cell(row=row, column=3).value = respuestas[pregunta_id]
      line = row

   # Llenar último campo con fecha, técnico y revisado por
   ws.cell(row=line, column=1).value = str("Fecha:   " + datetime.now().strftime("%d/%m/%Y") + f"    Técnico:   {str(session.get('tecnico', '?'))}" + "    Revisado por:    Andrés Herrera")

   # Generar nombre de archivo
   activo = session.get('activo_fijo', 'SinAF')
   propietario = session.get('propietario', 'SinPropietario').replace(' ', '-')
   cargo = session.get('cargo', 'SinCargo').replace(' ', '-')
   
   nombre_archivo = f"Activo {activo} Checklist {config['empresa']} {config['tipo']} {propietario} {cargo}.xlsx"
   # Sanitizar nombre de archivo para Windows
   safe_nombre = "".join(c if c not in r'\/:*?"<>|' else '_' for c in nombre_archivo)
   
   # Ruta del directorio de red compartido
   network_dir = r"\\172.16.1.22\checklist$\2025"
   
   try:
      # Intentar crear la carpeta en la red (si los permisos lo permiten)
      os.makedirs(network_dir, exist_ok=True)
      output_path = os.path.join(network_dir, safe_nombre)
      wb.save(output_path)
   except Exception:
      # Si falla (por falta de permisos), guarda localmente como respaldo
      local_dir = 'output'
      os.makedirs(local_dir, exist_ok=True)
      output_path = os.path.join(local_dir, safe_nombre)
      wb.save(output_path)
   
   return output_path


@app.route('/home')
def limpiar():
   session.clear()
   return redirect(url_for('index'))


if __name__ == '__main__':
   os.makedirs('static/img', exist_ok=True)
   os.makedirs('templates_excel', exist_ok=True)
   app.run(debug=True, host='0.0.0.0', port=9015)